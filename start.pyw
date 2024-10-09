import dash
from dash import dcc, html
import pandas as pd
from openpyxl import Workbook, load_workbook
import threading
import time
import os
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import webbrowser
from datetime import datetime, timedelta


def clean_old_data(df, max_points=360):  # 360 точек = 1 час при интервале в 10 секунд
    """Очищает старые данные, оставляя только последние max_points записей"""
    if len(df) > max_points:
        return df.iloc[-max_points:]
    return df


def load_data():
    try:
        wb = load_workbook('ground_motion.xlsx')
        ws = wb.active
        data = {
            'Timestamp': [],
            'Acceleration': [],
            'Velocity': [],
            'Displacement': []
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if isinstance(row[0], datetime):
                    timestamp = row[0]
                else:
                    if isinstance(row[0], (int, float)):
                        timestamp = pd.Timestamp.fromtimestamp(float(row[0]))
                    else:
                        timestamp = pd.to_datetime(row[0], errors='raise')
                acceleration = float(row[1]) if row[1] is not None else None
                velocity = float(row[2]) if row[2] is not None else None
                displacement = float(row[3]) if row[3] is not None else None
                data['Timestamp'].append(timestamp)
                data['Acceleration'].append(acceleration)
                data['Velocity'].append(velocity)
                data['Displacement'].append(displacement)
            except (ValueError, TypeError):
                continue

        df = pd.DataFrame(data)
        df = clean_old_data(df)  # Очищаем старые данные при загрузке
        return df
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Acceleration', 'Velocity', 'Displacement'])
        wb.save('ground_motion.xlsx')
        return pd.DataFrame(columns=['Timestamp', 'Acceleration', 'Velocity', 'Displacement'])
    except Exception:
        return pd.DataFrame(columns=['Timestamp', 'Acceleration', 'Velocity', 'Displacement'])


def save_data(df):
    """Сохраняет очищенные данные в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.append(['Timestamp', 'Acceleration', 'Velocity', 'Displacement'])

    for _, row in df.iterrows():
        ws.append([row['Timestamp'], row['Acceleration'], row['Velocity'], row['Displacement']])

    wb.save('ground_motion.xlsx')


def update_data():
    while True:
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                url = 'https://stationview.raspberryshake.org/#/?lat=43.13353&lon=76.90885&zoom=10.477&sta=RD437'
                page.goto(url)
                page.wait_for_selector('section.gm.col', timeout=10000)
                html = page.content()
                soup = BeautifulSoup(html, 'html.parser')
                section = soup.find('section', class_='gm col')
                if section:
                    values = section.find_all('span', class_='val')
                    if len(values) >= 3:
                        def extract_number(text):
                            try:
                                return float(''.join(c for c in text if c.isdigit() or c in '.-'))
                            except ValueError:
                                return None

                        timestamp = datetime.now()
                        acceleration = extract_number(values[0].text)
                        velocity = extract_number(values[1].text)
                        displacement = extract_number(values[2].text)

                        if all(v is not None for v in [acceleration, velocity, displacement]):
                            df = load_data()
                            new_data = pd.DataFrame({
                                'Timestamp': [timestamp],
                                'Acceleration': [acceleration],
                                'Velocity': [velocity],
                                'Displacement': [displacement]
                            })
                            df = pd.concat([df, new_data], ignore_index=True)
                            df = clean_old_data(df)
                            save_data(df)
                browser.close()
        except Exception:
            pass
        finally:
            time.sleep(10)


thread = threading.Thread(target=update_data)
thread.daemon = True
thread.start()

app = dash.Dash(__name__)

df = load_data()

app.layout = html.Div(style={'fontFamily': 'Arial', 'padding': '20px'}, children=[
    html.Button('Открыть файл Excel', id='open-excel-button', n_clicks=0, style={
        'backgroundColor': '#4CAF50', 'color': 'white', 'border': 'none', 'padding': '15px 32px',
        'textAlign': 'center', 'textDecoration': 'none', 'display': 'inline-block', 'fontSize': '16px',
        'margin': '4px 2px', 'cursor': 'pointer'
    }),
    html.H1(children='Данные о движении грунта', style={'textAlign': 'center', 'color': '#4CAF50'}),
    html.Div(id='data-display', style={'textAlign': 'center', 'margin': '20px 0'}),
    dcc.Graph(id='acceleration-graph'),
    dcc.Graph(id='velocity-graph'),
    dcc.Graph(id='displacement-graph'),
    dcc.Interval(
        id='interval-component',
        interval=60*1000,
        n_intervals=0
    )
])

@app.callback(
    dash.dependencies.Output('open-excel-button', 'n_clicks'),
    [dash.dependencies.Input('open-excel-button', 'n_clicks')]
)
def open_excel(n_clicks):
    if n_clicks > 0:
        webbrowser.open('file://' + os.path.realpath('ground_motion.xlsx'))
    return 0

@app.callback(
    [dash.dependencies.Output('data-display', 'children'),
     dash.dependencies.Output('acceleration-graph', 'figure'),
     dash.dependencies.Output('velocity-graph', 'figure'),
     dash.dependencies.Output('displacement-graph', 'figure')],
    [dash.dependencies.Input('interval-component', 'n_intervals')]
)
def update_graphs(n):
    df = load_data()
    if not df.empty:
        latest_data = df.iloc[-1]
        data_display = [
            html.P(f"Последние измерения движения грунта:"),
            html.P(f"Время: {latest_data['Timestamp'].strftime('%Y-%m-%d %H:%M:%S')}"),
            html.P(f"Ускорение: {latest_data['Acceleration']:.6f} м/с²"),
            html.P(f"Скорость: {latest_data['Velocity']:.6f} м/с"),
            html.P(f"Смещение: {latest_data['Displacement']:.6f} м")
        ]
        def create_figure(title, y_column, color):
            return {
                'data': [{'x': df['Timestamp'], 'y': df[y_column], 'type': 'line', 'name': title, 'line': {'color': color}}],
                'layout': {
                    'title': f'{title} с течением времени',
                    'xaxis': {'title': 'Время', 'tickformat': '%Y-%m-%d %H:%M:%S'},
                    'yaxis': {'title': f'{title} ({y_column})'},
                    'plot_bgcolor': '#f9f9f9',
                    'paper_bgcolor': '#f9f9f9',
                }
            }
        acceleration_figure = create_figure('Ускорение', 'Acceleration', 'red')
        velocity_figure = create_figure('Скорость', 'Velocity', 'blue')
        displacement_figure = create_figure('Смещение', 'Displacement', 'green')
        return data_display, acceleration_figure, velocity_figure, displacement_figure
    return "Нет данных", {}, {}, {}

if __name__ == '__main__':
    app.run_server(debug=True)